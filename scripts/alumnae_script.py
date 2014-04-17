from openpyxl import *
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

from pymongo import MongoClient

def main():

	client = MongoClient('mongodb://facsem:buffalotexas@troup.mongohq.com:10007/app22602359')
	db = client['app22602359']

	collection = db['unregistered']

	# print collection

	wb = load_workbook(filename = r'/Users/Clayton/Desktop/Alumnae Spreadsheet/From AlumnaeOffice_Computer Science Majors.xlsx')

	sheet_ranges = wb['Sheet1']

	# {}
	titleArray = []

	for row in xrange(1, 94):

		jsonString = {}

		for col_idx in xrange(1, 17):

			col = get_column_letter(col_idx)
			cellNum = str(col)+str(row)

			if row == 1:
				titleArray.append(sheet_ranges[cellNum].value)
			else:
				
				if isinstance(sheet_ranges[cellNum].value, unicode):
					jsonString[titleArray[col_idx-1].encode('ascii')] = sheet_ranges[cellNum].value.encode('ascii')
				# print type(sheet_ranges[cellNum].value)
				else:
					jsonString[titleArray[col_idx-1].encode('ascii')] = str(sheet_ranges[cellNum].value)

		collection.insert(jsonString)
		



if __name__ == '__main__':
	main()